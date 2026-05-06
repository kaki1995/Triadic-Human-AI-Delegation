# triadic_simulation/io_excel.py
from __future__ import annotations

from copy import copy
from typing import Any, Dict, List
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.worksheet.worksheet import Worksheet


def _clean_colname(x: object) -> str:
    """Normalize Excel header cell to a clean column name."""
    if x is None:
        return ""
    s = str(x).strip()
    # guard against common junk headers
    if s.lower() in {"none", "nan", "null"}:
        return ""
    return s


def safe_sheet_to_df(ws: Worksheet) -> pd.DataFrame:
    """
    Read a worksheet where the first row is a header.

    Robustness:
    - Trims header whitespace
    - Drops empty/None headers
    - Drops fully-empty rows
    """
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()

    raw_header = rows[0]
    header = [_clean_colname(x) for x in raw_header]
    header = [h for h in header if h]  # drop empty header cells

    if not header:
        return pd.DataFrame()

    # data rows: take only as many columns as header
    data: list[list[Any]] = []
    for r in rows[1:]:
        if not r:  # skip None / empty tuples
            continue
        rr = list(r)[: len(header)]
        # drop fully-empty rows
        if all(v is None for v in rr):
            continue
        data.append(rr)

    df = pd.DataFrame(data, columns=header)

    # drop accidental unnamed columns if any slipped through
    df = df.loc[:, [c for c in df.columns if c and not str(c).startswith("Unnamed")]]

    return df


def write_df_to_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    """
    Overwrite worksheet with DataFrame content.

    Robustness:
    - Clears entire sheet (all rows)
    - Writes header + rows
    - Coerces numpy scalars to Python types for openpyxl compatibility
    """
    # Clear sheet safely
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    # Ensure columns are strings
    cols = [str(c) for c in df.columns]
    ws.append(cols)

    # Write rows; convert numpy types to Python scalars
    for row in df.itertuples(index=False, name=None):
        out: list[Any] = []
        for v in row:
            if pd.isna(v):
                out.append(None)
            else:
                # pandas/numpy scalar -> python scalar
                if hasattr(v, "item"):
                    try:
                        out.append(v.item())
                        continue
                    except (TypeError, ValueError, AttributeError):
                        pass
                out.append(v)
        ws.append(out)


def _excel_value(v: Any) -> Any:
    """Convert pandas/numpy scalar values into Excel-safe Python values."""
    if pd.isna(v):
        return None
    if hasattr(v, "item"):
        try:
            return v.item()
        except (TypeError, ValueError, AttributeError):
            return v
    return v


def _append_df_write_only(ws: Any, df: pd.DataFrame) -> None:
    """Write a DataFrame to an openpyxl write-only worksheet."""
    ws.append([str(c) for c in df.columns])
    for row in df.itertuples(index=False, name=None):
        ws.append([_excel_value(v) for v in row])


def _copy_template_sheet_write_only(src_ws: Worksheet, dst_ws: Any) -> None:
    """Copy worksheet values into a write-only worksheet."""
    for row in src_ws.iter_rows(values_only=False):
        out = []
        for cell in row:
            out_cell = WriteOnlyCell(dst_ws, value=cell.value)
            if cell.has_style:
                out_cell.font = copy(cell.font)
                out_cell.fill = copy(cell.fill)
                out_cell.border = copy(cell.border)
                out_cell.alignment = copy(cell.alignment)
                out_cell.number_format = cell.number_format
            out.append(out_cell)
        dst_ws.append(out)


def ensure_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    """
    Ensure df has all columns in `cols`:
    - Add missing columns as NA
    - Drop extra columns not in `cols`
    - Return in the exact `cols` order
    """
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA

    return df[cols]


def write_to_schema_workbook(
    input_schema_xlsx: str,
    output_xlsx: str,
    dfs: Dict[str, pd.DataFrame],
    sheet_map: Dict[str, str],
) -> None:
    """
    Write DataFrames into an existing schema workbook.

    Behavior:
    - If sheet exists: align df columns to existing sheet header columns and overwrite
    - If sheet missing: create it and write df as-is

    Practical notes:
    - The schema workbook defines the "official" column order and acts as a template.
    - IMPORTANT: If the template sheet has headers, any df columns not present in the
      template WILL be dropped. This function prints a warning when that happens.
    """
    template_wb = load_workbook(input_schema_xlsx, read_only=False, data_only=False)
    output_wb = Workbook(write_only=True)
    written_sheets: set[str] = set()
    sheet_to_key = {sheet_name: df_key for df_key, sheet_name in sheet_map.items()}

    for df_key in sheet_map:
        if df_key not in dfs:
            raise KeyError(f"DataFrame key '{df_key}' missing from dfs. Available: {list(dfs.keys())}")

    for sheet_name in template_wb.sheetnames:
        src_ws = template_wb[sheet_name]
        if sheet_name not in sheet_to_key:
            dst_ws = output_wb.create_sheet(sheet_name)
            _copy_template_sheet_write_only(src_ws, dst_ws)
            written_sheets.add(sheet_name)
            continue

        df_key = sheet_to_key[sheet_name]
        df = dfs[df_key].copy()
        df.columns = [str(c).strip() for c in df.columns]

        existing = safe_sheet_to_df(src_ws)
        if not existing.empty and len(existing.columns) > 0:
            template_cols = list(existing.columns)
            extra = [c for c in df.columns if c not in template_cols]
            if extra:
                preview = extra[:8]
                suffix = "..." if len(extra) > 8 else ""
                print(f"WARNING [{df_key}] Dropping {len(extra)} columns not in template: {preview}{suffix}")
            df = ensure_columns(df, template_cols)

        dst_ws = output_wb.create_sheet(sheet_name)
        _append_df_write_only(dst_ws, df)
        written_sheets.add(sheet_name)

    for df_key, sheet_name in sheet_map.items():
        if sheet_name in written_sheets:
            continue
        df = dfs[df_key].copy()
        df.columns = [str(c).strip() for c in df.columns]
        dst_ws = output_wb.create_sheet(sheet_name)
        _append_df_write_only(dst_ws, df)

    output_wb.save(output_xlsx)
