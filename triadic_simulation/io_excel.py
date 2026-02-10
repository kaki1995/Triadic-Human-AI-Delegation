# triadic_sim/io_excel.py
from __future__ import annotations

from typing import Dict, List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _clean_colname(x: object) -> str:
    """Normalize Excel header cell to a clean column name."""
    if x is None:
        return ""
    s = str(x).strip()
    # guard against common junk headers
    if s.lower() in {"none", "nan", "null"}:
        return ""
    return s


def safe_sheet_to_df(ws: Worksheet) -> pd.DataFrame:
    """
    Read a worksheet where the first row is a header.

    Robustness:
    - Trims header whitespace
    - Drops empty/None headers
    - Drops fully-empty rows
    """
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()

    raw_header = rows[0]
    header = [_clean_colname(x) for x in raw_header]
    header = [h for h in header if h]  # drop empty header cells

    if not header:
        return pd.DataFrame()

    # data rows: take only as many columns as header
    data = []
    for r in rows[1:]:
        if r is None:
            continue
        rr = list(r)[: len(header)]
        # drop fully-empty rows
        if all(v is None for v in rr):
            continue
        data.append(rr)

    df = pd.DataFrame(data, columns=header)

    # drop accidental unnamed columns if any slipped through
    df = df.loc[:, [c for c in df.columns if c and not str(c).startswith("Unnamed")]]

    return df


def write_df_to_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    """
    Overwrite worksheet with DataFrame content.

    Robustness:
    - Clears entire sheet (all rows)
    - Writes header + rows
    - Coerces numpy scalars to Python types for openpyxl compatibility
    """
    # Clear sheet safely
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    # Ensure columns are strings
    cols = [str(c) for c in df.columns]
    ws.append(cols)

    # Write rows; convert numpy types to Python scalars
    for row in df.itertuples(index=False, name=None):
        out = []
        for v in row:
            if pd.isna(v):
                out.append(None)
            else:
                # pandas/numpy scalar -> python scalar
                if hasattr(v, "item"):
                    try:
                        out.append(v.item())
                        continue
                    except Exception:
                        pass
                out.append(v)
        ws.append(out)


def ensure_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    """
    Ensure df has all columns in `cols`:
    - Add missing columns as NA
    - Drop extra columns not in `cols`
    - Return in the exact `cols` order
    """
    # normalize df columns to strings (avoid mismatches)
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA

    # keep only schema columns, in schema order
    return df[cols]


def write_to_schema_workbook(
    input_schema_xlsx: str,
    output_xlsx: str,
    dfs: Dict[str, pd.DataFrame],
    sheet_map: Dict[str, str],
) -> None:
    """
    Write DataFrames into an existing schema workbook.

    Behavior:
    - If sheet exists: align df columns to existing sheet header columns and overwrite
    - If sheet missing: create it and write df as-is

    Practical notes:
    - This supports your workflow where the schema workbook defines the "official"
      column order and acts as a template.
    """
    wb = load_workbook(input_schema_xlsx)

    for df_key, sheet_name in sheet_map.items():
        if df_key not in dfs:
            raise KeyError(f"DataFrame key '{df_key}' missing from dfs. Available: {list(dfs.keys())}")

        df = dfs[df_key].copy()

        # Normalize df columns (helps alignment with Excel template)
        df.columns = [str(c).strip() for c in df.columns]

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            existing = safe_sheet_to_df(ws)

            # If template has headers, enforce them (order + missing cols)
            if not existing.empty and len(existing.columns) > 0:
                df = ensure_columns(df, list(existing.columns))

            write_df_to_sheet(ws, df)
        else:
            ws = wb.create_sheet(sheet_name)
            write_df_to_sheet(ws, df)

    wb.save(output_xlsx)
